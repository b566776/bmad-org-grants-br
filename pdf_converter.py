#!/usr/bin/env python3
"""
Módulo unificado para conversão de arquivos para Markdown.

- PDF -> Markdown: Docling (primário) e pypdf (fallback)
- XLSX -> Markdown: openpyxl (modelos de orçamento/anexos do edital)
"""

import os
import json
import sys
from pathlib import Path
from typing import Optional, Dict, Any, Tuple, List
from datetime import datetime

# Tentar importar Docling
DOCLING_AVAILABLE = False
try:
    from docling.document_converter import DocumentConverter
    DOCLING_AVAILABLE = True
except ImportError:
    DOCLING_AVAILABLE = False

# Tentar importar pypdf
PYPDF_AVAILABLE = False
try:
    import pypdf
    PYPDF_AVAILABLE = True
except ImportError:
    PYPDF_AVAILABLE = False

# Tentar importar openpyxl (XLSX)
OPENPYXL_AVAILABLE = False
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def load_config() -> Dict[str, Any]:
    """Carrega configuração do arquivo config.json"""
    config_path = Path("config/config.json")
    default_config = {
        "pdf_conversion": {
            "default_engine": "auto",
            "fallback_to_pypdf": True,
            "docling_options": {
                "enable_ocr": True,
                "preserve_tables": True,
                "preserve_formulas": True
            }
        }
    }
    
    if config_path.exists():
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
                # Mescla com configuração padrão
                if "pdf_conversion" not in config:
                    config["pdf_conversion"] = default_config["pdf_conversion"]
                return config
        except Exception as e:
            print(f"Aviso: Erro ao carregar config.json: {e}", file=sys.stderr)
    
    return default_config


def detect_available_engine(preferred: Optional[str] = None) -> Tuple[str, bool]:
    """
    Detecta qual engine está disponível
    
    Args:
        preferred: Engine preferido ('docling', 'pypdf', 'auto')
        
    Returns:
        Tupla (engine_name, is_available)
    """
    config = load_config()
    default_engine = config.get("pdf_conversion", {}).get("default_engine", "auto")
    
    # Usa preferência explícita ou configuração padrão
    engine_pref = preferred or default_engine
    
    if engine_pref == "docling":
        return ("docling", DOCLING_AVAILABLE)
    elif engine_pref == "pypdf":
        return ("pypdf", PYPDF_AVAILABLE)
    else:  # auto
        # Prioriza Docling se disponível
        if DOCLING_AVAILABLE:
            return ("docling", True)
        elif PYPDF_AVAILABLE:
            return ("pypdf", True)
        else:
            return ("none", False)


def find_pdf_file(pdf_path: str) -> Path:
    """
    Encontra o arquivo PDF usando múltiplas estratégias
    
    Args:
        pdf_path: Caminho para o arquivo PDF
        
    Returns:
        Path do arquivo encontrado
    """
    path = Path(pdf_path)
    workspace = Path.cwd()
    
    # 1. Se for absoluto e existir
    if path.is_absolute() and path.exists():
        return path
    
    # 2. Tenta relativo ao workspace
    test_path = workspace / path
    if test_path.exists():
        return test_path
    
    # 3. Busca recursiva em memories/ (específico para este módulo)
    memories_dir = workspace / "memories"
    if memories_dir.exists():
        for root, dirs, files in os.walk(memories_dir):
            test_path = Path(root) / path.name
            if test_path.exists() and test_path.suffix.lower() == '.pdf':
                return test_path
    
    # 4. Busca recursiva limitada no workspace
    for root, dirs, files in os.walk(workspace):
        if len(Path(root).relative_to(workspace).parts) > 3:
            continue
        test_path = Path(root) / path.name
        if test_path.exists() and test_path.suffix.lower() == '.pdf':
            return test_path
    
    # Retorna o caminho original (vai gerar erro descritivo)
    return path


def find_xlsx_file(xlsx_path: str) -> Path:
    """
    Encontra o arquivo XLSX usando múltiplas estratégias.

    Observação: o módulo só suporta XLSX (não XLS).
    """
    path = Path(xlsx_path)
    workspace = Path.cwd()

    # 1. Se for absoluto e existir
    if path.is_absolute() and path.exists():
        return path

    # 2. Tenta relativo ao workspace
    test_path = workspace / path
    if test_path.exists():
        return test_path

    # 3. Busca recursiva em memories/ (específico para este módulo)
    memories_dir = workspace / "memories"
    if memories_dir.exists():
        for root, dirs, files in os.walk(memories_dir):
            test_path = Path(root) / path.name
            if test_path.exists() and test_path.suffix.lower() == ".xlsx":
                return test_path

    # 4. Busca recursiva limitada no workspace
    for root, dirs, files in os.walk(workspace):
        if len(Path(root).relative_to(workspace).parts) > 3:
            continue
        test_path = Path(root) / path.name
        if test_path.exists() and test_path.suffix.lower() == ".xlsx":
            return test_path

    return path


def _is_row_empty(row: List[str]) -> bool:
    return all((c or "").strip() == "" for c in row)


def _trim_empty_edges(matrix: List[List[str]]) -> List[List[str]]:
    """
    Remove linhas e colunas totalmente vazias nas bordas.
    Não tenta otimização perfeita: é uma heurística robusta para planilhas modelo.
    """
    if not matrix:
        return []

    # remove linhas vazias no topo
    top = 0
    while top < len(matrix) and _is_row_empty(matrix[top]):
        top += 1

    # remove linhas vazias no fim
    bottom = len(matrix) - 1
    while bottom >= top and _is_row_empty(matrix[bottom]):
        bottom -= 1

    if bottom < top:
        return []

    trimmed = matrix[top : bottom + 1]

    # descobrir colunas não vazias
    max_cols = max((len(r) for r in trimmed), default=0)
    if max_cols == 0:
        return []

    # normaliza tamanhos
    normalized: List[List[str]] = [r + [""] * (max_cols - len(r)) for r in trimmed]

    left = 0
    while left < max_cols and all((row[left] or "").strip() == "" for row in normalized):
        left += 1

    right = max_cols - 1
    while right >= left and all((row[right] or "").strip() == "" for row in normalized):
        right -= 1

    if right < left:
        return []

    return [row[left : right + 1] for row in normalized]


def _markdown_escape_cell(value: str) -> str:
    # evita quebrar tabela
    return (value or "").replace("|", "\\|").replace("\n", " ").strip()


def _sheet_to_markdown_table(
    matrix: List[List[str]],
    max_rows: int,
    max_cols: int,
) -> Tuple[str, Dict[str, Any]]:
    """
    Renderiza uma matriz 2D em tabela Markdown.
    A primeira linha é usada como header.
    """
    meta: Dict[str, Any] = {"truncated": False, "rows": 0, "cols": 0}

    if not matrix:
        return "_(aba vazia)_\n", meta

    # trim bordas vazias
    matrix = _trim_empty_edges(matrix)
    if not matrix:
        return "_(aba vazia)_\n", meta

    # aplica limites
    rows = matrix[:max_rows]
    cols_len = max((len(r) for r in rows), default=0)
    cols_len = min(cols_len, max_cols)
    rows = [r[:cols_len] + [""] * (cols_len - len(r[:cols_len])) for r in rows]

    meta["rows"] = len(rows)
    meta["cols"] = cols_len
    if len(matrix) > max_rows or any(len(r) > max_cols for r in matrix):
        meta["truncated"] = True

    header = rows[0]
    body = rows[1:] if len(rows) > 1 else []

    header_cells = [_markdown_escape_cell(c) or " " for c in header]
    sep_cells = ["---"] * cols_len

    lines: List[str] = []
    lines.append("| " + " | ".join(header_cells) + " |")
    lines.append("| " + " | ".join(sep_cells) + " |")
    for r in body:
        cells = [_markdown_escape_cell(c) for c in r]
        lines.append("| " + " | ".join(cells) + " |")

    return "\n".join(lines) + "\n", meta


def convert_excel_to_markdown(
    xlsx_path: str,
    output_path: Optional[str] = None,
    verbose: bool = False,
    max_rows: int = 200,
    max_cols: int = 30,
) -> Tuple[str, Dict[str, Any]]:
    """
    Converte XLSX para Markdown (1 .md por workbook, com seções por aba).

    - O .md é salvo no mesmo diretório e com o mesmo nome do .xlsx, por padrão.
    - Não suporta .xls.
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl não está disponível. Instale com: pip install openpyxl")

    xlsx_file = find_xlsx_file(xlsx_path)
    if not xlsx_file.exists():
        raise FileNotFoundError(f"XLSX não encontrado: {xlsx_path}")
    if xlsx_file.suffix.lower() != ".xlsx":
        raise ValueError(f"Arquivo não é um XLSX: {xlsx_path}")

    out_path = xlsx_file.with_suffix(".md") if output_path is None else Path(output_path)

    if verbose:
        print(f"📊 Convertendo XLSX: {xlsx_file}", file=sys.stderr)

    wb = openpyxl.load_workbook(filename=str(xlsx_file), data_only=True, read_only=True)

    md_parts: List[str] = []
    metadata: Dict[str, Any] = {
        "engine": "openpyxl",
        "file_path": str(xlsx_file.absolute()),
        "file_name": xlsx_file.name,
        "conversion_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "sheets": [],
        "max_rows": max_rows,
        "max_cols": max_cols,
    }

    for ws in wb.worksheets:
        # ignora abas ocultas
        if getattr(ws, "sheet_state", "visible") != "visible":
            continue

        sheet_name = str(ws.title)
        md_parts.append(f"## Aba: {sheet_name}\n")

        matrix: List[List[str]] = []
        # read_only: ws.iter_rows é eficiente
        # coletamos como string (para markdown)
        for row in ws.iter_rows(values_only=True):
            matrix.append(["" if v is None else str(v) for v in row])

        table_md, table_meta = _sheet_to_markdown_table(matrix, max_rows=max_rows, max_cols=max_cols)
        md_parts.append(table_md)

        if table_meta.get("truncated"):
            md_parts.append(
                f"> ⚠️ Tabela truncada para {max_rows} linhas e {max_cols} colunas.\n\n"
            )
        else:
            md_parts.append("\n")

        metadata["sheets"].append(
            {
                "name": sheet_name,
                "rows": table_meta.get("rows", 0),
                "cols": table_meta.get("cols", 0),
                "truncated": bool(table_meta.get("truncated", False)),
            }
        )

    # fecha workbook
    try:
        wb.close()
    except Exception:
        pass

    # se nenhuma aba visível produziu conteúdo
    if not md_parts:
        md_parts.append("_(nenhuma aba visível encontrada)_\n")

    content = "\n".join(md_parts).strip() + "\n"
    enhanced = enhance_markdown_metadata(content, metadata)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(enhanced, encoding="utf-8")

    if verbose:
        print(f"✅ XLSX convertido para: {out_path}", file=sys.stderr)

    return str(out_path), metadata


def convert_file_to_markdown(
    path: str,
    output_path: Optional[str] = None,
    engine: Optional[str] = None,
    fallback: bool = True,
    verbose: bool = False,
) -> Tuple[str, Dict[str, Any]]:
    """
    Dispatcher: converte PDF ou XLSX para Markdown conforme extensão.
    Mantém `convert_pdf_to_markdown` intacto para compatibilidade.
    """
    p = Path(path)
    suffix = p.suffix.lower()
    if suffix == ".pdf":
        return convert_pdf_to_markdown(
            pdf_path=path,
            output_path=output_path,
            engine=engine,
            fallback=fallback,
            verbose=verbose,
        )
    if suffix == ".xlsx":
        return convert_excel_to_markdown(
            xlsx_path=path,
            output_path=output_path,
            verbose=verbose,
        )
    raise ValueError(f"Extensão não suportada para conversão: {suffix}")


def convert_with_docling(pdf_path: Path, output_path: Optional[Path] = None) -> Tuple[str, Dict[str, Any]]:
    """
    Converte PDF para Markdown usando Docling
    
    Args:
        pdf_path: Caminho para o arquivo PDF
        output_path: Caminho de saída (opcional)
        
    Returns:
        Tupla (markdown_content, metadata)
    """
    if not DOCLING_AVAILABLE:
        raise ImportError("Docling não está disponível. Instale com: pip install docling")
    
    config = load_config()
    docling_options = config.get("pdf_conversion", {}).get("docling_options", {})
    
    # Cria o conversor
    converter = DocumentConverter()
    
    # Converte o documento
    result = converter.convert(str(pdf_path))
    
    # Extrai Markdown
    markdown_content = result.document.export_to_markdown()
    
    # Extrai metadados
    metadata = {
        "engine": "docling",
        "file_path": str(pdf_path.absolute()),
        "file_name": pdf_path.name,
        "conversion_date": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }
    
    # Adiciona metadados do documento se disponíveis
    if hasattr(result.document, 'metadata') and result.document.metadata:
        doc_metadata = result.document.metadata
        if hasattr(doc_metadata, 'title'):
            metadata["title"] = str(doc_metadata.title)
        if hasattr(doc_metadata, 'author'):
            metadata["author"] = str(doc_metadata.author)
    
    return markdown_content, metadata


def convert_with_pypdf(pdf_path: Path, output_path: Optional[Path] = None) -> Tuple[str, Dict[str, Any]]:
    """
    Converte PDF para Markdown usando pypdf (fallback)
    
    Args:
        pdf_path: Caminho para o arquivo PDF
        output_path: Caminho de saída (opcional)
        
    Returns:
        Tupla (markdown_content, metadata)
    """
    if not PYPDF_AVAILABLE:
        raise ImportError("pypdf não está disponível. Instale com: pip install pypdf")
    
    text_parts = []
    metadata = {
        "engine": "pypdf",
        "file_path": str(pdf_path.absolute()),
        "file_name": pdf_path.name,
        "conversion_date": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }
    
    with open(pdf_path, 'rb') as file:
        pdf_reader = pypdf.PdfReader(file)
        
        # Extrair metadados do PDF
        if pdf_reader.metadata:
            pdf_metadata = pdf_reader.metadata
            if pdf_metadata.get("/Title"):
                metadata["title"] = str(pdf_metadata.get("/Title", ""))
            if pdf_metadata.get("/Author"):
                metadata["author"] = str(pdf_metadata.get("/Author", ""))
            if pdf_metadata.get("/Subject"):
                metadata["subject"] = str(pdf_metadata.get("/Subject", ""))
        
        # Extrair texto de todas as páginas
        total_pages = len(pdf_reader.pages)
        metadata["total_pages"] = total_pages
        
        for page_num, page in enumerate(pdf_reader.pages, 1):
            page_text = page.extract_text()
            if page_text.strip():
                # Limpa quebras de linha excessivas
                cleaned_text = '\n'.join(line.strip() for line in page_text.split('\n') if line.strip())
                text_parts.append(f"## Página {page_num}\n\n{cleaned_text}\n")
    
    markdown_content = '\n'.join(text_parts)
    
    return markdown_content, metadata


def enhance_markdown_metadata(markdown_content: str, metadata: Dict[str, Any]) -> str:
    """
    Adiciona cabeçalho com metadados ao Markdown
    
    Args:
        markdown_content: Conteúdo Markdown
        metadata: Metadados do documento
        
    Returns:
        Markdown com cabeçalho de metadados
    """
    header_parts = []
    
    # Título
    title = metadata.get("title") or metadata.get("file_name", "Documento")
    title = title.replace(".pdf", "").replace(".PDF", "").replace(".xlsx", "").replace(".XLSX", "")
    header_parts.append(f"# {title}\n")
    
    # Metadados
    if metadata.get("author"):
        header_parts.append(f"**Autor:** {metadata['author']}\n")
    if metadata.get("subject"):
        header_parts.append(f"**Assunto:** {metadata['subject']}\n")
    
    header_parts.append(f"\n**Arquivo original:** `{metadata.get('file_name', 'N/A')}`\n")
    
    if metadata.get("total_pages"):
        header_parts.append(f"**Total de páginas:** {metadata['total_pages']}\n")
    
    header_parts.append(f"**Engine de conversão:** {metadata.get('engine', 'N/A')}\n")
    header_parts.append(f"**Data de conversão:** {metadata.get('conversion_date', 'N/A')}\n")
    header_parts.append("\n---\n\n")
    
    return '\n'.join(header_parts) + markdown_content


def convert_pdf_to_markdown(
    pdf_path: str,
    output_path: Optional[str] = None,
    engine: Optional[str] = None,
    fallback: bool = True,
    verbose: bool = False
) -> Tuple[str, Dict[str, Any]]:
    """
    Função principal unificada para conversão de PDF para Markdown
    
    Args:
        pdf_path: Caminho para o arquivo PDF
        output_path: Caminho de saída (opcional, usa mesmo diretório do PDF se não especificado)
        engine: Engine a usar ('docling', 'pypdf', 'auto')
        fallback: Se True, tenta fallback automático em caso de erro
        verbose: Se True, imprime informações sobre o processo
        
    Returns:
        Tupla (caminho_do_arquivo_md, metadata)
    """
    # Encontra o arquivo PDF
    pdf_file = find_pdf_file(pdf_path)
    
    if not pdf_file.exists():
        raise FileNotFoundError(f"PDF não encontrado: {pdf_path}")
    
    if pdf_file.suffix.lower() != '.pdf':
        raise ValueError(f"Arquivo não é um PDF: {pdf_path}")
    
    # Define caminho de saída
    if output_path is None:
        output_path = pdf_file.with_suffix('.md')
    else:
        output_path = Path(output_path)
    
    # Detecta engine disponível
    engine_name, is_available = detect_available_engine(engine)
    
    if not is_available:
        if fallback:
            # Tenta o outro engine
            if engine_name == "docling":
                engine_name, is_available = detect_available_engine("pypdf")
            else:
                engine_name, is_available = detect_available_engine("docling")
        
        if not is_available:
            raise RuntimeError(
                "Nenhum engine de conversão disponível. "
                "Instale pelo menos um: pip install docling OU pip install pypdf"
            )
    
    # Carrega configuração
    config = load_config()
    fallback_enabled = config.get("pdf_conversion", {}).get("fallback_to_pypdf", True)
    if fallback is not None:
        fallback_enabled = fallback
    
    markdown_content = None
    metadata = {}
    used_engine = engine_name
    fallback_used = False
    
    # Tenta conversão com engine escolhido
    try:
        if verbose:
            print(f"Usando engine: {engine_name}", file=sys.stderr)
        
        if engine_name == "docling":
            markdown_content, metadata = convert_with_docling(pdf_file)
        elif engine_name == "pypdf":
            markdown_content, metadata = convert_with_pypdf(pdf_file)
        else:
            raise ValueError(f"Engine desconhecido: {engine_name}")
            
    except Exception as e:
        if verbose:
            print(f"Erro com {engine_name}: {e}", file=sys.stderr)
        
        # Tenta fallback se habilitado
        if fallback_enabled and engine_name != "pypdf":
            try:
                if verbose:
                    print("Tentando fallback para pypdf...", file=sys.stderr)
                markdown_content, metadata = convert_with_pypdf(pdf_file)
                used_engine = "pypdf"
                fallback_used = True
            except Exception as fallback_error:
                raise RuntimeError(
                    f"Falha na conversão com {engine_name} e fallback para pypdf também falhou. "
                    f"Erros: {e}; {fallback_error}"
                )
        else:
            raise
    
    # Adiciona informação sobre fallback nos metadados
    metadata["engine_used"] = used_engine
    metadata["fallback_used"] = fallback_used
    
    # Melhora o Markdown com metadados
    enhanced_markdown = enhance_markdown_metadata(markdown_content, metadata)
    
    # Salva o arquivo
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(enhanced_markdown, encoding='utf-8')
    
    if verbose:
        print(f"✅ Conversão concluída usando {used_engine}", file=sys.stderr)
        if fallback_used:
            print(f"⚠️  Fallback usado: {engine_name} → pypdf", file=sys.stderr)
        print(f"📄 Arquivo salvo em: {output_path}", file=sys.stderr)
    
    return str(output_path), metadata
